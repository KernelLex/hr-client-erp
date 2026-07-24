"""
Profitability, Opex, Inventory, Ageing (enhanced), and Transport endpoints
for the Accounting Dashboard (Cards 1–5).

Shared utility: get_profitability_summary(from_date, to_date) is importable
so other modules can call it without re-computing stock values.

Data sources:
  - VE Tally Voucher / VE Tally Ledger / VE Tally Stock Item  (Tally import)
  - VE Stock Movement Summary                                  (accounts_tally_import)
  - Vera Expense Claim                                         (manual entry)
  - VE Transport Record                                        (manual entry, new)
"""
import datetime
import json
import frappe
from frappe.utils import flt, cint, today

from hr_client.api.utils import require_admin, handle_api_error


# ── Period helpers ─────────────────────────────────────────────────────────────

def _period_bounds(period: str, custom_start: str = None, custom_end: str = None):
    t = datetime.date.today()
    if period == "today":
        d = t.isoformat(); return d, d
    if period == "mtd":
        return t.replace(day=1).isoformat(), t.isoformat()
    if period == "ytd":
        fy_year = t.year if t.month >= 4 else t.year - 1
        return f"{fy_year}-04-01", t.isoformat()
    if period == "last_month":
        first_this = t.replace(day=1)
        last_prev = first_this - datetime.timedelta(days=1)
        return last_prev.replace(day=1).isoformat(), last_prev.isoformat()
    if period == "custom":
        return (custom_start or t.replace(day=1).isoformat()), (custom_end or t.isoformat())
    return t.replace(day=1).isoformat(), t.isoformat()


def _prior_period(from_date: str, to_date: str):
    s = datetime.date.fromisoformat(from_date)
    e = datetime.date.fromisoformat(to_date)
    span = (e - s).days + 1
    prev_e = s - datetime.timedelta(days=1)
    prev_s = prev_e - datetime.timedelta(days=span - 1)
    return prev_s.isoformat(), prev_e.isoformat()


def _pct(curr, prev):
    if not prev: return None
    return round((flt(curr) - flt(prev)) / abs(flt(prev)) * 100, 1)


# ── Shared: stock value snapshot ───────────────────────────────────────────────

def _stock_value_snapshot():
    """
    Returns (total_value, neg_value) using VE Stock Movement Summary × VE Tally Stock Item.
    This is a current-state snapshot — Tally doesn't expose per-period stock levels.
    """
    rows = frappe.db.sql("""
        SELECT
            COALESCE(SUM(GREATEST(s.stock_on_hand, 0) * COALESCE(i.standard_rate, 0)), 0) AS total_value,
            COALESCE(SUM(CASE WHEN s.stock_on_hand < 0
                         THEN ABS(s.stock_on_hand) * COALESCE(i.standard_rate, 0)
                         ELSE 0 END), 0) AS neg_value,
            COUNT(CASE WHEN s.stock_on_hand != 0 THEN 1 END) AS active_skus,
            COUNT(CASE WHEN s.stock_on_hand < 0 THEN 1 END) AS neg_sku_count
        FROM `tabVE Stock Movement Summary` s
        LEFT JOIN `tabVE Tally Stock Item` i ON i.item_name = s.item_code
    """, as_dict=True)
    r = rows[0] if rows else {}
    return (flt(r.get("total_value")), flt(r.get("neg_value")),
            cint(r.get("active_skus")), cint(r.get("neg_sku_count")))


# ── Shared: voucher totals ─────────────────────────────────────────────────────

def _voucher_total(vtypes: tuple, from_date: str, to_date: str) -> float:
    if not vtypes: return 0.0
    rows = frappe.db.sql(
        """SELECT COALESCE(SUM(amount), 0) AS total
           FROM `tabVE Tally Voucher`
           WHERE is_cancelled = 0 AND voucher_type IN %s
           AND voucher_date BETWEEN %s AND %s""",
        (vtypes, from_date, to_date), as_dict=True,
    )
    return flt(rows[0].total) if rows else 0.0


# ── Shared: opex computation ───────────────────────────────────────────────────

def _compute_opex(from_date: str, to_date: str):
    """
    Aggregate opex from three sources:
    1. Tally Ledger expense groups (YTD balance, best available without line-level Tally data)
    2. Vera Expense Claims approved in period (period-specific)
    3. VE Transport Records in period (period-specific)
    Returns (total_opex, breakdown_dict)
    """
    breakdown = {}

    # Source 1 — Tally expense ledger balances (YTD, not period-filtered)
    # Dr balance (negative) on expense accounts = expense incurred
    exp_rows = frappe.db.sql("""
        SELECT ledger_name, parent_group, closing_balance
        FROM `tabVE Tally Ledger`
        WHERE (
            parent_group LIKE '%Expense%'
            OR parent_group LIKE '%Salary%'
            OR parent_group LIKE '%Wages%'
            OR parent_group LIKE '%Rent%'
            OR parent_group LIKE '%Admin%'
        )
        AND closing_balance < 0
        ORDER BY closing_balance ASC
    """, as_dict=True)

    tally_cats = {}
    tally_total = 0.0
    for r in exp_rows:
        amt = abs(flt(r.closing_balance))
        tally_total += amt
        pg = (r.parent_group or "Other").strip()
        tally_cats[pg] = round(tally_cats.get(pg, 0) + amt, 2)

    if tally_total > 0:
        breakdown["Tally Ledger (YTD)"] = {"total": round(tally_total, 2), "categories": tally_cats}

    # Source 2 — Vera Expense Claims
    claim_total = 0.0
    try:
        claim_rows = frappe.db.sql(
            """SELECT claim_type, COALESCE(SUM(amount), 0) AS total
               FROM `tabVera Expense Claim`
               WHERE status = 'Approved' AND claim_date BETWEEN %s AND %s
               GROUP BY claim_type""",
            (from_date, to_date), as_dict=True,
        )
        claim_total = sum(flt(r.total) for r in claim_rows)
        if claim_total > 0:
            breakdown["Expense Claims"] = {
                "total": round(claim_total, 2),
                "categories": {r.claim_type: round(flt(r.total), 2) for r in claim_rows},
            }
    except Exception:
        pass

    # Source 3 — VE Transport Records
    transport_total = 0.0
    try:
        tr_rows = frappe.db.sql(
            """SELECT source, COALESCE(SUM(amount), 0) AS total
               FROM `tabVE Transport Record`
               WHERE entry_date BETWEEN %s AND %s
               GROUP BY source""",
            (from_date, to_date), as_dict=True,
        )
        transport_total = sum(flt(r.total) for r in tr_rows)
        if transport_total > 0:
            breakdown["Transport & Labour"] = {
                "total": round(transport_total, 2),
                "categories": {r.source: round(flt(r.total), 2) for r in tr_rows},
            }
    except Exception:
        pass

    opex_period = claim_total + transport_total   # period-specific
    opex_yd = tally_total                          # YTD Tally balance
    return round(opex_period, 2), round(opex_yd, 2), breakdown


# ── Public utility: get_profitability_summary ──────────────────────────────────

def get_profitability_summary(from_date: str, to_date: str):
    """
    Shared computation used by Cards 1 and 4.
    COGS = Purchases in period (proxy — Tally doesn't expose period-specific stock levels).
    For full COGS accounting, closing_stock value is shown separately.
    """
    gross_sales = _voucher_total(("Sales",), from_date, to_date)
    credit_notes = _voucher_total(("Credit Note",), from_date, to_date)
    net_sales = gross_sales - credit_notes

    purchases = _voucher_total(("Purchase",), from_date, to_date)
    debit_notes = _voucher_total(("Debit Note",), from_date, to_date)
    net_purchases = purchases - debit_notes

    closing_stock, neg_stock_value, active_skus, _ = _stock_value_snapshot()
    cogs = net_purchases   # simplified: stock change approximated as 0 for short periods
    gross_profit = net_sales - cogs

    opex_period, opex_ytd, opex_breakdown = _compute_opex(from_date, to_date)
    net_profit = gross_profit - opex_period

    prev_from, prev_to = _prior_period(from_date, to_date)
    prev_net_sales = _voucher_total(("Sales",), prev_from, prev_to) - _voucher_total(("Credit Note",), prev_from, prev_to)
    prev_purchases = _voucher_total(("Purchase",), prev_from, prev_to) - _voucher_total(("Debit Note",), prev_from, prev_to)
    prev_gross_profit = prev_net_sales - prev_purchases

    return {
        "from_date": from_date,
        "to_date": to_date,
        "net_sales": round(net_sales, 2),
        "gross_sales": round(gross_sales, 2),
        "credit_notes": round(credit_notes, 2),
        "net_purchases": round(net_purchases, 2),
        "cogs": round(cogs, 2),
        "closing_stock": round(closing_stock, 2),
        "gross_profit": round(gross_profit, 2),
        "opex_period": opex_period,
        "opex_ytd": opex_ytd,
        "opex_breakdown": opex_breakdown,
        "net_profit": round(net_profit, 2),
        "gross_margin_pct": round(gross_profit / net_sales * 100, 1) if net_sales else None,
        "net_margin_pct": round(net_profit / net_sales * 100, 1) if net_sales else None,
        "pop": {
            "net_sales_pct": _pct(net_sales, prev_net_sales),
            "gross_profit_pct": _pct(gross_profit, prev_gross_profit),
            "prev_from": prev_from,
            "prev_to": prev_to,
        },
    }


# ── Card 1: Profitability Summary ──────────────────────────────────────────────

@frappe.whitelist()
@handle_api_error
def get_card_profitability(period="mtd", custom_start=None, custom_end=None):
    require_admin()
    from_date, to_date = _period_bounds(period, custom_start, custom_end)
    return get_profitability_summary(from_date, to_date)


# ── Card 2: Enhanced Ageing Analysis ──────────────────────────────────────────

@frappe.whitelist()
@handle_api_error
def get_card_ageing():
    """
    Two-part ageing:
    A) Creditors & Debtors by invoice due date vs today
       Buckets: 0-20, 21-45, 46-90, 90+ days
    B) Advances to/from by advance date vs today
       Buckets: 0-6m, 7-12m, 13-24m, 24+m
    Drill-down lists included per bucket.
    """
    require_admin()
    today_str = today()

    def _invoice_ageing(doctype, date_col, amount_col, party_col):
        rows = frappe.db.sql(
            f"""SELECT {party_col} AS party, {amount_col} AS amount, {date_col} AS inv_date,
                       DATEDIFF(%s, {date_col}) AS days
                FROM `tab{doctype}`
                WHERE status NOT IN ('Cleared')""",
            (today_str,), as_dict=True,
        )
        buckets = {"b0_20": [], "b21_45": [], "b46_90": [], "b90plus": []}
        totals = {k: 0.0 for k in buckets}
        for r in rows:
            d = cint(r.days) if r.days is not None else 0
            amt = flt(r.amount)
            if d <= 20:    k = "b0_20"
            elif d <= 45:  k = "b21_45"
            elif d <= 90:  k = "b46_90"
            else:          k = "b90plus"
            buckets[k].append({"party": r.party, "amount": amt, "date": str(r.inv_date or ""), "days": d})
            totals[k] += amt
        return {
            "totals": {k: round(v, 2) for k, v in totals.items()},
            "grand_total": round(sum(totals.values()), 2),
            "rows": rows and [{"party": r.party, "amount": flt(r.amount), "date": str(r.inv_date or ""), "days": cint(r.days) if r.days is not None else 0} for r in rows],
        }

    def _advance_ageing(doctype, date_col, amount_col, party_col):
        rows = frappe.db.sql(
            f"""SELECT {party_col} AS party, {amount_col} AS amount, {date_col} AS adv_date,
                       TIMESTAMPDIFF(MONTH, {date_col}, %s) AS months
                FROM `tab{doctype}`""",
            (today_str,), as_dict=True,
        )
        buckets = {"b0_6m": [], "b7_12m": [], "b13_24m": [], "b24plus": []}
        totals = {k: 0.0 for k in buckets}
        for r in rows:
            m = cint(r.months) if r.months is not None else 0
            amt = flt(r.amount)
            if m <= 6:     k = "b0_6m"
            elif m <= 12:  k = "b7_12m"
            elif m <= 24:  k = "b13_24m"
            else:          k = "b24plus"
            buckets[k].append({"party": r.party, "amount": amt, "date": str(r.adv_date or ""), "months": m})
            totals[k] += amt
        return {
            "totals": {k: round(v, 2) for k, v in totals.items()},
            "grand_total": round(sum(totals.values()), 2),
            "rows": [{"party": r.party, "amount": flt(r.amount), "date": str(r.adv_date or ""), "months": cint(r.months) if r.months is not None else 0} for r in rows],
        }

    return {
        "as_of": today_str,
        "creditors":        _invoice_ageing("VE Creditor Ledger", "invoice_date", "due_amount", "vendor_name"),
        "debtors":          _invoice_ageing("VE Debtor Ledger", "invoice_date", "due_amount", "client_name"),
        "adv_to_creditors": _advance_ageing("VE Creditor Advance", "advance_date", "advance_amount", "vendor_name"),
        "adv_from_debtors": _advance_ageing("VE Debtor Advance", "advance_date", "advance_amount", "client_name"),
        "bucket_labels": {
            "invoice": {"b0_20": "0–20 days", "b21_45": "21–45 days", "b46_90": "46–90 days", "b90plus": "90+ days"},
            "advance": {"b0_6m": "0–6 months", "b7_12m": "7–12 months", "b13_24m": "13–24 months", "b24plus": "24+ months"},
        },
    }


# ── Card 3: Inventory Summary ──────────────────────────────────────────────────

@frappe.whitelist()
@handle_api_error
def get_card_inventory():
    require_admin()

    total_val, neg_val, active_skus, neg_sku_count = _stock_value_snapshot()

    # Groups / brands breakdown
    group_rows = frappe.db.sql("""
        SELECT i.stock_group AS grp,
               COUNT(DISTINCT s.item_code) AS sku_count,
               COALESCE(SUM(GREATEST(s.stock_on_hand, 0) * COALESCE(i.standard_rate, 0)), 0) AS value
        FROM `tabVE Stock Movement Summary` s
        LEFT JOIN `tabVE Tally Stock Item` i ON i.item_name = s.item_code
        WHERE i.stock_group IS NOT NULL AND i.stock_group != ''
        GROUP BY i.stock_group
        ORDER BY value DESC
        LIMIT 30
    """, as_dict=True)

    # Category / movement type breakdown
    cat_rows = frappe.db.sql("""
        SELECT s.movement_category AS category,
               COUNT(*) AS sku_count,
               COALESCE(SUM(GREATEST(s.stock_on_hand, 0) * COALESCE(i.standard_rate, 0)), 0) AS value
        FROM `tabVE Stock Movement Summary` s
        LEFT JOIN `tabVE Tally Stock Item` i ON i.item_name = s.item_code
        GROUP BY s.movement_category
        ORDER BY value DESC
    """, as_dict=True)

    total_skus = frappe.db.count("VE Stock Movement Summary")
    reorder_count = frappe.db.count("VE Stock Movement Summary", {"movement_category": "Reorder"})

    return {
        "total_stock_value": round(total_val, 2),
        "negative_stock_value": round(neg_val, 2),
        "negative_stock_value_display": round(abs(neg_val), 2),
        "active_skus": active_skus,
        "negative_sku_count": neg_sku_count,
        "total_sku_count": total_skus,
        "reorder_alert_count": reorder_count,
        "by_group": [
            {"group": r.grp, "sku_count": cint(r.sku_count), "value": round(flt(r.value), 2)}
            for r in group_rows
        ],
        "by_category": [
            {"category": r.category, "sku_count": cint(r.sku_count), "value": round(flt(r.value), 2)}
            for r in cat_rows
        ],
    }


# ── Card 4: Opex / Admin Expenses ─────────────────────────────────────────────

@frappe.whitelist()
@handle_api_error
def get_card_opex(period="mtd", custom_start=None, custom_end=None):
    require_admin()
    from_date, to_date = _period_bounds(period, custom_start, custom_end)
    # Reuse shared profitability computation
    summary = get_profitability_summary(from_date, to_date)
    return {
        "from_date": from_date,
        "to_date": to_date,
        "opex_period": summary["opex_period"],
        "opex_ytd": summary["opex_ytd"],
        "breakdown": summary["opex_breakdown"],
    }


# ── Card 5: Transport & Labour ─────────────────────────────────────────────────

@frappe.whitelist()
@handle_api_error
def get_card_transport(period="mtd", custom_start=None, custom_end=None):
    require_admin()
    from_date, to_date = _period_bounds(period, custom_start, custom_end)

    rows = frappe.db.sql(
        """SELECT source,
                  COALESCE(SUM(amount), 0) AS total,
                  COALESCE(SUM(labour_day_charges), 0) AS day_charges,
                  COALESCE(SUM(labour_transport), 0) AS labour_transport,
                  COALESCE(SUM(labour_food), 0) AS labour_food,
                  COUNT(*) AS entry_count
           FROM `tabVE Transport Record`
           WHERE entry_date BETWEEN %s AND %s
           GROUP BY source""",
        (from_date, to_date), as_dict=True,
    )

    by_source = {
        r.source: {
            "total": round(flt(r.total), 2),
            "entry_count": cint(r.entry_count),
            "day_charges": round(flt(r.day_charges), 2),
            "labour_transport": round(flt(r.labour_transport), 2),
            "labour_food": round(flt(r.labour_food), 2),
        }
        for r in rows
    }

    total = sum(v["total"] for v in by_source.values())

    recent = frappe.db.sql(
        """SELECT name, source, entry_date, amount, description,
                  labour_day_charges, labour_transport, labour_food, notes
           FROM `tabVE Transport Record`
           WHERE entry_date BETWEEN %s AND %s
           ORDER BY entry_date DESC LIMIT 50""",
        (from_date, to_date), as_dict=True,
    )

    return {
        "from_date": from_date,
        "to_date": to_date,
        "total": round(total, 2),
        "by_source": by_source,
        "recent": [dict(r) for r in recent],
    }


@frappe.whitelist()
@handle_api_error
def create_transport_record(
    source, entry_date, amount=None,
    description=None, notes=None,
    labour_day_charges=None, labour_transport=None, labour_food=None,
):
    require_admin()
    if source not in ("Porter", "Rapido", "Other", "Labour"):
        frappe.throw("Invalid source. Must be Porter, Rapido, Other, or Labour.")

    doc = frappe.new_doc("VE Transport Record")
    doc.source = source
    doc.entry_date = entry_date
    doc.description = description or ""
    doc.notes = notes or ""

    if source == "Labour":
        doc.labour_day_charges = flt(labour_day_charges)
        doc.labour_transport = flt(labour_transport)
        doc.labour_food = flt(labour_food)
        doc.amount = flt(amount) if amount else (
            doc.labour_day_charges + doc.labour_transport + doc.labour_food
        )
    else:
        doc.amount = flt(amount)

    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"success": True, "name": doc.name}


@frappe.whitelist()
@handle_api_error
def upload_transport_csv():
    """
    Parse an uploaded CSV/Excel file and bulk-create VE Transport Record entries.
    Expected columns: source, entry_date, amount, description, notes
    For Labour rows: additionally labour_day_charges, labour_transport, labour_food
    """
    require_admin()
    import io
    import csv

    uploaded_file = frappe.request.files.get("file")
    source_override = frappe.form_dict.get("source")   # optional: force all rows to this source

    if not uploaded_file:
        frappe.throw("No file uploaded.")

    filename = uploaded_file.filename.lower()
    content = uploaded_file.read()

    rows_created = 0
    errors = []

    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        data_rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            data_rows = [
                {headers[i]: (str(cell.value or "").strip()) for i, cell in enumerate(row)}
                for row in ws.iter_rows(min_row=2)
                if any(cell.value for cell in row)
            ]
        except ImportError:
            frappe.throw("openpyxl not installed. Upload a CSV file instead.")
    else:
        frappe.throw("Unsupported file type. Upload .csv or .xlsx")

    for i, row in enumerate(data_rows, start=2):
        try:
            src = source_override or row.get("source", "").strip()
            if src not in ("Porter", "Rapido", "Other", "Labour"):
                errors.append(f"Row {i}: invalid source '{src}'")
                continue

            entry_date = str(row.get("entry_date", "")).strip()
            if not entry_date:
                errors.append(f"Row {i}: missing entry_date")
                continue

            doc = frappe.new_doc("VE Transport Record")
            doc.source = src
            doc.entry_date = entry_date
            doc.description = row.get("description", "")[:140]
            doc.notes = row.get("notes", "")[:200]
            doc.reference_doc = filename

            if src == "Labour":
                doc.labour_day_charges = flt(row.get("labour_day_charges") or row.get("day_charges") or 0)
                doc.labour_transport   = flt(row.get("labour_transport") or 0)
                doc.labour_food        = flt(row.get("labour_food") or row.get("food") or 0)
                doc.amount = flt(row.get("amount") or 0) or (
                    doc.labour_day_charges + doc.labour_transport + doc.labour_food
                )
            else:
                doc.amount = flt(row.get("amount") or 0)

            doc.insert(ignore_permissions=True)
            rows_created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    frappe.db.commit()
    return {
        "success": True,
        "rows_created": rows_created,
        "errors": errors,
        "message": f"Created {rows_created} records." + (f" {len(errors)} rows skipped." if errors else ""),
    }


@frappe.whitelist()
@handle_api_error
def sync_porter_api():
    """
    Stub for future Porter API integration.
    Wire up the real Porter credentials and endpoint here when ready.
    """
    require_admin()
    frappe.log_error("sync_porter_api called but not yet configured", "Porter API Stub")
    return {
        "success": False,
        "message": "Porter API integration not yet configured. "
                   "Add Porter API credentials to site config and implement sync_porter_api() in profitability.py.",
    }
